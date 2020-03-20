#https://www.youtube.com/watch?v=_uQrJ0TkZlc

#==========================

import openpyxl as xl

wb = xl.load_workbook('Files/transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
sheet.cell = (1, 1)
print(cell.value)
print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)